from io import BytesIO
from datetime import date
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import BoxConfig

# ── Per-component metadata: (werkstoff, festigkeit, oberfläche) ──────────────

_ITEM_META = {
    "y_corner":            ("S235 JRH", "-",   "verzinkt"),
    "l_corner":            ("S235 JRH", "-",   "verzinkt"),
    "t_corner":            ("S235 JRH", "-",   "verzinkt"),
    "k_corner":            ("S235 JRH", "-",   "verzinkt"),
    "tube_100x100":        ("S235 JRH", "-",   "verzinkt"),
    "roof_substructure":   ("S235 JR",  "-",   "verzinkt"),
    "roof_panel":          ("S235 JR",  "-",   "verzinkt"),
    "base_plate":          ("S235 JRH", "-",   "verzinkt"),
    "bolt_m12_120_a2":     ("Stahl",    "8.8", "verzinkt"),
    "nut_m12":             ("Stahl",    "8",   "verzinkt"),
    "schiene":             ("Stahl",    "-",   "verzinkt"),
    "wall_panel":          ("-",        "-",   "-"),
    "floor_panel":         ("-",        "-",   "-"),
    "bike_stand_rail":     ("S235 JRH", "-",   "verzinkt"),
    "bike_stand":          ("S235 JRH", "-",   "verzinkt"),
    "roller_door":         ("Stahl",    "-",   "-"),
    "solar_panel":         ("Aluminium / Glas", "-", "-"),
    "crossbeam":           ("Stahl",    "-",   "verzinkt"),
    "glass_clamp_profile": ("Aluminium","-",   "eloxiert"),
    "gutter_section":      ("Aluminium","-",   "-"),
    "gutter_connector":    ("Aluminium","-",   "-"),
    "gutter_end_cap":      ("Aluminium","-",   "-"),
    "downpipe":            ("Aluminium","-",   "-"),
}

_NORM_MAP = {
    "bolt_m12_120_a2": "DIN 933",
    "nut_m12":         "EN ISO 7040",
}

_FRAME_COLOR_NAMES = {
    "tiefschwarz":      "RAL 9005 Tiefschwarz",
    "verkehrsweiss":    "RAL 9016 Verkehrsweiß",
    "anthrazitgrau":    "RAL 7016 Anthrazitgrau",
    "lichtgrau":        "RAL 7035 Lichtgrau",
    "feuerrot":         "RAL 3000 Feuerrot",
    "enzianblau":       "RAL 5010 Enzianblau",
    "moosgruen":        "RAL 6005 Moosgrün",
    "schokoladenbraun": "RAL 8017 Schokoladenbraun",
}

# ── Colours & fonts ───────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_ALT_FILL  = PatternFill("solid", fgColor="D9E1F2")
_WHITE     = PatternFill("solid", fgColor="FFFFFF")
_HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
_TITLE_FONT= Font(name="Calibri", bold=True, size=12)
_META_BOLD = Font(name="Calibri", bold=True, size=9)
_META_NORM = Font(name="Calibri", size=9)
_DATA_FONT = Font(name="Calibri", size=9)
_WRAP      = Alignment(wrap_text=True, vertical="top")
_WRAP_C    = Alignment(wrap_text=True, vertical="top", horizontal="center")
_WRAP_R    = Alignment(wrap_text=True, vertical="top", horizontal="right")

_THIN = Side(style="thin", color="B0B0B0")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=Side(style="medium", color="1F3864"))


def _set(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:          cell.font = font
    if fill:          cell.fill = fill
    if alignment:     cell.alignment = alignment
    if border:        cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def generate_bom_xlsx(config: BoxConfig, bom_items: List[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stückliste"

    L, W, H = config.width_mm, config.depth_mm, config.height_mm
    fc_val = config.frame_color.value if config.frame_color else None
    fc_name = _FRAME_COLOR_NAMES.get(fc_val, "verzinkt") if fc_val else "verzinkt"
    steel_surface = f"verzinkt, {fc_name}" if fc_val else "verzinkt"

    today = date.today().strftime("%d.%m.%Y")
    drawing_nr = f"AB-{date.today().strftime('%Y%m%d')}"
    title_str = f"TER-BOX AB1000 – {L:.0f}×{W:.0f}×{H:.0f} mm"

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {1: 5, 2: 10, 3: 7, 4: 16, 5: 34, 6: 15, 7: 20, 8: 14, 9: 12, 10: 22}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    _set(ws, 1, 2, "Materialstückliste", font=_TITLE_FONT, alignment=_WRAP)

    # ── Rows 2-7: metadata block ──────────────────────────────────────────────
    for r in range(2, 8):
        ws.row_dimensions[r].height = 15

    meta_left = [
        ("Titel",           title_str,   "C2", "F2"),
        ("Zeichnungs-Nr.:", drawing_nr,  "C3", "F3"),
        ("Revision:",       "1.0",       "C4", "F4"),
        ("Abmessung:",      None,        None, None),
    ]

    _set(ws, 2, 2, "Titel",            font=_META_BOLD, alignment=_WRAP)
    _set(ws, 2, 3, title_str,          font=_META_NORM, alignment=_WRAP)
    _set(ws, 3, 2, "Zeichnungs-Nr.:",  font=_META_BOLD, alignment=_WRAP)
    _set(ws, 3, 3, drawing_nr,         font=_META_NORM, alignment=_WRAP)
    _set(ws, 4, 2, "Revision:",        font=_META_BOLD, alignment=_WRAP)
    _set(ws, 4, 3, "1.0",              font=_META_NORM, alignment=_WRAP)
    _set(ws, 5, 2, "Abmessung:",       font=_META_BOLD, alignment=_WRAP)

    _set(ws, 5, 3, "Länge:",  font=_META_BOLD, alignment=_WRAP)
    _set(ws, 5, 4, L,         font=_META_NORM, alignment=_WRAP)
    _set(ws, 5, 5, "mm",      font=_META_NORM, alignment=_WRAP)
    _set(ws, 6, 3, "Breite:", font=_META_BOLD, alignment=_WRAP)
    _set(ws, 6, 4, W,         font=_META_NORM, alignment=_WRAP)
    _set(ws, 6, 5, "mm",      font=_META_NORM, alignment=_WRAP)
    _set(ws, 7, 3, "Höhe:",   font=_META_BOLD, alignment=_WRAP)
    _set(ws, 7, 4, H,         font=_META_NORM, alignment=_WRAP)
    _set(ws, 7, 5, "mm",      font=_META_NORM, alignment=_WRAP)

    _set(ws, 2, 8, "Erstellt von:",  font=_META_BOLD, alignment=_WRAP)
    _set(ws, 2, 9, "TER-BOX System", font=_META_NORM, alignment=_WRAP)
    _set(ws, 3, 8, "Datum:",         font=_META_BOLD, alignment=_WRAP)
    _set(ws, 3, 9, today,            font=_META_NORM, alignment=_WRAP)
    _set(ws, 4, 8, "Geprüft von:",   font=_META_BOLD, alignment=_WRAP)
    _set(ws, 4, 9, "Terhardt, Sven", font=_META_NORM, alignment=_WRAP)

    ws.merge_cells("B5:B7")
    ws.merge_cells("C2:G2")
    ws.merge_cells("C3:G3")
    ws.merge_cells("C4:G4")
    ws.merge_cells("I2:J2")
    ws.merge_cells("I3:J3")
    ws.merge_cells("I4:J4")

    # ── Row 8: empty separator ────────────────────────────────────────────────
    ws.row_dimensions[8].height = 6

    # ── Row 9: column headers ─────────────────────────────────────────────────
    ws.row_dimensions[9].height = 30
    headers = ["Pos.", "Anzahl", "Einheit", "Zeich.-Nr.", "Benennung",
               "Norm", "Abmessung", "Werkstoff", "Festigkeits-\nklasse", "Oberfläche"]
    for c, h in enumerate(headers, 1):
        _set(ws, 9, c, h, font=_HDR_FONT, fill=_HDR_FILL,
             alignment=_WRAP_C, border=_HDR_BORDER)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for pos_idx, item in enumerate(bom_items, 1):
        row = 9 + pos_idx
        ws.row_dimensions[row].height = 15
        fill = _ALT_FILL if pos_idx % 2 == 0 else _WHITE

        key    = item.get("component_key", "")
        meta   = _ITEM_META.get(key, ("-", "-", "-"))
        werkst, festig, oberfl_default = meta

        # Surface: steel structural parts get frame color
        if oberfl_default == "verzinkt":
            oberfl = steel_surface
        else:
            oberfl = oberfl_default

        # For wall/floor panels use note as surface description
        if key in ("wall_panel", "floor_panel") and item.get("note"):
            oberfl = item["note"].split(",")[0] if "," in item["note"] else item["note"]

        # Abmessung column
        abmessung = "-"
        if item.get("length_mm"):
            abmessung = f"{int(item['length_mm'])} mm"
        elif item.get("note"):
            # Strip the leading description from notes like "Längsträger oben 2600 mm"
            note = item["note"]
            # If note ends with "mm" extract just the dimension part
            parts = note.rsplit(" ", 2)
            if len(parts) >= 2 and parts[-1] == "mm":
                abmessung = f"{parts[-2]} mm"
            else:
                abmessung = note

        norm = _NORM_MAP.get(key, "-")

        cells = [
            (1, pos_idx,                  _DATA_FONT, _WRAP_C),
            (2, item.get("qty", 1),       _DATA_FONT, _WRAP_C),
            (3, "Stk.",                   _DATA_FONT, _WRAP_C),
            (4, item.get("article_nr", "-"), _DATA_FONT, _WRAP),
            (5, item.get("description", "-"), _DATA_FONT, _WRAP),
            (6, norm,                     _DATA_FONT, _WRAP_C),
            (7, abmessung,                _DATA_FONT, _WRAP),
            (8, werkst,                   _DATA_FONT, _WRAP_C),
            (9, festig,                   _DATA_FONT, _WRAP_C),
            (10, oberfl,                  _DATA_FONT, _WRAP),
        ]
        for col, val, font, align in cells:
            _set(ws, row, col, val, font=font, fill=fill, alignment=align, border=_THIN_BORDER)

    # ── Freeze header rows ────────────────────────────────────────────────────
    ws.freeze_panes = "A10"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = f"A9:J{9 + len(bom_items)}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
